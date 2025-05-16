#!/usr/bin/env python3
"""
JSON to Excel Converter for High/Low Order Text Data

This script processes JSON data containing high-order and low-order text relationships,
and generates an Excel file formatted according to a specific template.
"""

import json
import os
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Convert JSON data to Excel format.')
    parser.add_argument('--input', '-i', required=True, help='Input JSON file path')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    return parser.parse_args()

def load_json_data(file_path):
    """Load and validate JSON data from file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return data
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON format in {file_path}")
        exit(1)
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        exit(1)

def process_json_data(json_data):
    """
    Process JSON data to extract and format high-order and low-order text relationships.
    The data is extracted exactly as is without any modifications.
    
    Returns a list of dictionaries ready for DataFrame creation.
    """
    rows = []
    
    # Track publication IDs that have already had reasoning added (for reasoning assignment)
    used_publication_ids = set()
    
    # Support multiple JSON root structures
    # 1. If input is a dict with 'payload'->'results', use that
    if isinstance(json_data, dict) and 'payload' in json_data and 'results' in json_data['payload']:
        items = json_data['payload']['results']
    # 2. If it's a list, use as is
    elif isinstance(json_data, list):
        items = json_data
    # 3. If it's a dict (single item), wrap in list
    elif isinstance(json_data, dict):
        items = [json_data]
    else:
        print("Error: Unsupported JSON structure.")
        exit(1)
    # Process each item in the JSON data
    for item in items:
        # Extract high-order text information
        high_order_texts = item.get('high_order_text', [])
        reasonings = item.get('reasonings', [])
        
        # Process each high-order text entry
        for hot in high_order_texts:
            # Create row for high-order text
            high_order_row = {
                'Text Type': 'High-Order Text',
                'Paragraph ID': hot.get('paragraph_ID', ''),
                'Publication ID': hot.get('publication_ID', ''),
                'Task Text': hot.get('text', ''),
                'Tag': ', '.join(hot.get('tags', [])),
                'Similarity Score': 'N/A',
                'Reasonings': ''
            }
            rows.append(high_order_row)
            
            # Process each associated low-order text
            for lot in hot.get('low_order_texts', []):
                publication_id = lot.get('publication_ID', '')
                paragraph_id = lot.get('paragraph_ID', '')
                tag = lot.get('tag', f"INCON-{hot.get('paragraph_ID', '')}")
                similarity_score = lot.get('similarity_score', '')
                task_text = lot.get('text', '')

                # Assign reasoning only for the first occurrence of each publication_ID (case sensitive)
                reasoning_text = None
                if publication_id and publication_id not in used_publication_ids:
                    for reasoning in reasonings:
                        # Use correct key: publication_ID (case sensitive)
                        if reasoning.get('publication_ID') == publication_id:
                            reasoning_text = reasoning.get('reasoning', None)
                            break
                    used_publication_ids.add(publication_id)

                # Handle CONF- tags: set all fields except Tag and Reasonings to None
                if tag and str(tag).startswith("CONF-"):
                    low_order_row = {
                        'Text Type': None,
                        'Paragraph ID': None,
                        'Publication ID': None,
                        'Task Text': None,
                        'Tag': tag,
                        'Similarity Score': None,
                        'Reasonings': reasoning_text
                    }
                    rows.append(low_order_row)
                    continue

                low_order_row = {
                    'Text Type': 'Low-Order Text',
                    'Paragraph ID': paragraph_id,
                    'Publication ID': publication_id,
                    'Task Text': task_text,
                    'Tag': tag,
                    'Similarity Score': similarity_score,
                    'Reasonings': reasoning_text
                }
                rows.append(low_order_row)
    
    return rows

def create_excel_file(data_rows, output_path):
    """
    Create an Excel file from the processed data.
    
    Args:
        data_rows: List of dictionaries containing row data
        output_path: Path for the output Excel file
    """
    # Create DataFrame from processed data
    df = pd.DataFrame(data_rows)
    
    # Write DataFrame to Excel without index
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Apply formatting to match template
    format_excel_file(output_path)
    
    print(f"Excel file created successfully at: {output_path}")

def format_excel_file(file_path):
    """
    Apply formatting to the Excel file to match the template.
    
    Args:
        file_path: Path to the Excel file to format
    """
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
    
    # Format data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the formatted workbook
    wb.save(file_path)

def main():
    """Main execution function."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Load JSON data
    json_data = load_json_data(args.input)
    
    # Process JSON data
    processed_data = process_json_data(json_data)
    
    # Create Excel file
    create_excel_file(processed_data, args.output)

if __name__ == "__main__":
    main()